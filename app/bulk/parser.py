import csv
import os
from typing import Iterable, List, Tuple

import openpyxl
import xlrd


def iter_rows(file_path: str) -> Tuple[List[str], Iterable[List[str]]]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext in {".xlsx"}:
        return _iter_xlsx(file_path)
    if ext in {".csv"}:
        return _iter_csv(file_path)
    if ext in {".xls"}:
        return _iter_xls(file_path)
    raise ValueError("Formato nao suportado. Use XLSX ou CSV.")


def _is_instruction_row(row: List[str]) -> bool:
    if not row:
        return False
    keywords = [
        "obrigatorio",
        "opcional",
        "separar",
        "sim/nao",
        "sim nao",
        "padrao",
        "itens",
    ]
    hits = 0
    for cell in row:
        raw = (cell or "").strip().lower()
        if not raw:
            hits += 1
            continue
        if any(key in raw for key in keywords):
            hits += 1
    return hits >= max(1, int(len(row) * 0.6))


def _iter_xlsx(file_path: str) -> Tuple[List[str], Iterable[List[str]]]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]

    def _gen():
        pending = next(rows, None)
        if pending is not None:
            pending_row = ["" if c is None else str(c).strip() for c in pending]
            if not _is_instruction_row(pending_row):
                yield pending_row
        for row in rows:
            yield ["" if c is None else str(c).strip() for c in row]

    return header, _gen()


def _iter_csv(file_path: str) -> Tuple[List[str], Iterable[List[str]]]:
    def _gen():
        with open(file_path, newline="", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            header = next(reader)
            first = next(reader, None)
            if first is not None:
                first_row = [cell.strip() for cell in first]
                if not _is_instruction_row(first_row):
                    yield first_row
            for row in reader:
                yield [cell.strip() for cell in row]

    with open(file_path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        header = next(reader)
    return [h.strip() for h in header], _gen()


def _iter_xls(file_path: str) -> Tuple[List[str], Iterable[List[str]]]:
    book = xlrd.open_workbook(file_path)
    sheet = book.sheet_by_index(0)
    header = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]

    def _gen():
        if sheet.nrows > 1:
            first_row = [str(sheet.cell_value(1, col)).strip() for col in range(sheet.ncols)]
            if not _is_instruction_row(first_row):
                yield first_row
            start = 2
        else:
            start = 1
        for row_idx in range(start, sheet.nrows):
            row = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
            yield row

    return header, _gen()
