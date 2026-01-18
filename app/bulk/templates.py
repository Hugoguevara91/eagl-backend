from datetime import datetime
from io import BytesIO
from typing import Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.bulk.config import ENTITY_CONFIGS
from app.db import models


def build_template(entity: str) -> Tuple[bytes, str]:
    config = ENTITY_CONFIGS[entity]
    wb = Workbook()
    ws = wb.active
    ws.title = "MODELO"

    headers = [col.label for col in config.template_columns]
    instructions = [col.instruction for col in config.template_columns]
    ws.append(headers)
    ws.append(instructions)

    ws.freeze_panes = "A3"
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 26

    info = wb.create_sheet("INFO")
    info["A1"] = "Template"
    info["B1"] = entity
    info["A2"] = "Versao"
    info["B2"] = config.template_version
    info["A3"] = "Gerado em"
    info["B3"] = datetime.utcnow().isoformat()

    out = BytesIO()
    wb.save(out)
    filename = f"modelo_{entity}_{config.template_version}.xlsx"
    return out.getvalue(), filename


def build_export(db, tenant_id: str, entity: str) -> Tuple[bytes, str]:
    config = ENTITY_CONFIGS[entity]
    wb = Workbook()
    ws = wb.active
    ws.title = "EXPORT"
    headers = [col.label for col in config.template_columns]
    ws.append(headers)

    if entity == "employees":
        items = (
            db.query(models.Colaborador)
            .filter(models.Colaborador.tenant_id == tenant_id)
            .order_by(models.Colaborador.created_at.desc())
            .all()
        )
        for item in items:
            ws.append(
                [
                    item.nome,
                    item.funcao,
                    item.email,
                    item.telefone,
                    item.status,
                    item.contrato,
                    item.unidade,
                    item.coordenador_nome,
                    item.supervisor_nome,
                    ";".join(item.especialidades or []),
                    item.observacoes,
                ]
            )
    elif entity == "clients":
        items = (
            db.query(models.Client)
            .filter(models.Client.tenant_id == tenant_id)
            .order_by(models.Client.created_at.desc())
            .all()
        )
        for item in items:
            ws.append(
                [
                    item.name,
                    item.client_code,
                    item.document,
                    item.status,
                    item.contract,
                    item.address,
                ]
            )
    elif entity == "sites":
        items = (
            db.query(models.Site, models.CustomerAccount)
            .outerjoin(models.CustomerAccount, models.CustomerAccount.id == models.Site.customer_account_id)
            .filter(models.Site.tenant_id == tenant_id)
            .order_by(models.Site.created_at.desc())
            .all()
        )
        for site, account in items:
            ws.append(
                [
                    site.code,
                    site.name,
                    account.cnpj if account else None,
                    account.name if account else None,
                    site.status,
                    site.address,
                ]
            )
    elif entity == "assets":
        items = (
            db.query(models.Asset, models.Client, models.Site)
            .outerjoin(models.Client, models.Client.id == models.Asset.client_id)
            .outerjoin(models.Site, models.Site.id == models.Asset.site_id)
            .filter(models.Asset.tenant_id == tenant_id)
            .order_by(models.Asset.created_at.desc())
            .all()
        )
        for asset, client, site in items:
            ws.append(
                [
                    asset.tag,
                    asset.name,
                    asset.asset_type,
                    asset.status,
                    client.document if client else None,
                    client.client_code if client else None,
                    site.code if site else None,
                ]
            )
    elif entity == "os_types":
        items = (
            db.query(models.OSType, models.Client)
            .outerjoin(models.Client, models.Client.id == models.OSType.client_id)
            .filter(models.OSType.tenant_id == tenant_id)
            .order_by(models.OSType.created_at.desc())
            .all()
        )
        for item, client in items:
            ws.append(
                [
                    item.name,
                    item.description,
                    client.document if client else None,
                    client.client_code if client else None,
                    "SIM" if item.is_active else "NAO",
                ]
            )
    elif entity == "questionnaires":
        items = (
            db.query(models.Questionnaire, models.QuestionnaireItem)
            .join(models.QuestionnaireItem, models.QuestionnaireItem.questionnaire_id == models.Questionnaire.id)
            .filter(models.Questionnaire.tenant_id == tenant_id)
            .order_by(models.Questionnaire.created_at.desc())
            .all()
        )
        for questionnaire, q_item in items:
            ws.append(
                [
                    questionnaire.title,
                    questionnaire.version,
                    q_item.question_text,
                    "SIM" if q_item.required else "NAO",
                    q_item.answer_type,
                    ";".join(q_item.items or []),
                ]
            )

    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 26

    out = BytesIO()
    wb.save(out)
    filename = f"export_{entity}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx"
    return out.getvalue(), filename
