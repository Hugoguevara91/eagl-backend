import hashlib
import os
import re
import tempfile
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook

from app.bulk.config import ENTITY_CONFIGS, label_for_key, make_header_map, normalize_header
from app.bulk.parser import iter_rows
from app.bulk.storage import StorageClient
from app.db import models


MAX_PREVIEW_ROWS = 20
CHUNK_SIZE = 500


class ImportValidationError(Exception):
    pass


def _resolve_unique_key(row: Dict[str, object], unique_key_groups: List[List[str]]) -> Tuple[str, ...] | None:
    for group in unique_key_groups:
        if all(row.get(key) not in (None, "") for key in group):
            return tuple(str(row.get(key)).strip() for key in group)
    return None


def _append_error(errors: List[dict], row_number: int, field: str, message: str, severity: str = "error"):
    errors.append(
        {"row_number": row_number, "field": field, "message": message, "severity": severity}
    )

def _is_valid_email(value: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value))


def _is_valid_cnpj(value: str) -> bool:
    digits = "".join(ch for ch in value if ch.isdigit())
    return len(digits) == 14


def _should_skip(mode: str, exists: bool) -> bool:
    if mode == "create_only" and exists:
        return True
    if mode == "update_only" and not exists:
        return True
    return False


def validate_job(db, job: models.ImportJob) -> dict:
    config = ENTITY_CONFIGS.get(job.entity)
    if not config:
        raise ImportValidationError("Entidade nao suportada.")

    job.status = "validating"
    db.commit()

    storage = StorageClient()
    local_path = storage.download_to_temp(job.file_url)
    header, rows = iter_rows(local_path)
    header_map = make_header_map(config.template_columns)
    canonical_headers = [header_map.get(normalize_header(h), "") for h in header]

    required = [c.key for c in config.template_columns if c.required]
    missing_required = [col for col in required if col not in canonical_headers]
    if missing_required:
        raise ImportValidationError(
            "Faltou a coluna " + ", ".join(label_for_key(config, col).upper() for col in missing_required)
        )

    errors: List[dict] = []
    preview = {"created": 0, "updated": 0, "skipped": 0, "errors": 0, "samples": []}
    seen_keys = set()

    db.query(models.ImportRowError).filter(models.ImportRowError.import_job_id == job.id).delete()

    for idx, row in enumerate(rows, start=3):
        row_data: Dict[str, object] = {}
        for key, value in zip(canonical_headers, row):
            if not key:
                continue
            raw = value.strip() if isinstance(value, str) else value
            if raw in (None, ""):
                continue
            transformer = config.transformers.get(key)
            try:
                row_data[key] = transformer(str(raw)) if transformer else raw
            except Exception:
                _append_error(errors, idx, key, f"Valor invalido para {label_for_key(config, key)}")

        if not any(str(cell).strip() for cell in row):
            continue

        for col in required:
            if col not in row_data or row_data.get(col) in (None, ""):
                _append_error(errors, idx, col, "Campo obrigatorio")

        if "email" in row_data and not _is_valid_email(str(row_data.get("email"))):
            _append_error(errors, idx, "email", "E-mail invalido")

        if row_data.get("document") and not _is_valid_cnpj(str(row_data.get("document"))):
            _append_error(errors, idx, "document", "CNPJ invalido")
        if row_data.get("customer_account_cnpj") and not _is_valid_cnpj(str(row_data.get("customer_account_cnpj"))):
            _append_error(errors, idx, "customer_account_cnpj", "CNPJ invalido")
        if row_data.get("client_cnpj") and not _is_valid_cnpj(str(row_data.get("client_cnpj"))):
            _append_error(errors, idx, "client_cnpj", "CNPJ invalido")

        unique_key = _resolve_unique_key(row_data, config.unique_key_groups)
        if not unique_key:
            _append_error(errors, idx, "__unique__", "Chave unica nao encontrada")
        elif job.entity != "questionnaires":
            if unique_key in seen_keys:
                _append_error(errors, idx, "__unique__", "Chave unica duplicada no arquivo")
            else:
                seen_keys.add(unique_key)

        if job.entity == "questionnaires":
            if "version" not in row_data:
                row_data["version"] = 1
            if row_data.get("answer_type", "").upper() == "ITENS" and not row_data.get("items"):
                _append_error(errors, idx, "items", "Itens obrigatorios quando Tipo=ITENS")

        if job.entity == "clients":
            if not row_data.get("document") and not row_data.get("client_code"):
                _append_error(errors, idx, "document", "Informe CNPJ ou Codigo do cliente")

        if job.entity == "sites":
            account_id = _resolve_customer_account_id(db, job.tenant_id, row_data)
            if not account_id:
                _append_error(errors, idx, "customer_account_cnpj", "Cliente nao encontrado pelo CNPJ")

        if job.entity == "assets":
            if row_data.get("client_cnpj") or row_data.get("client_code"):
                if not _resolve_client_id(db, job.tenant_id, row_data):
                    _append_error(errors, idx, "client_cnpj", "Cliente nao encontrado pelo CNPJ/Codigo")
            if row_data.get("site_code"):
                site = (
                    db.query(models.Site)
                    .filter(models.Site.tenant_id == job.tenant_id, models.Site.code == row_data.get("site_code"))
                    .first()
                )
                if not site:
                    _append_error(errors, idx, "site_code", "Site nao encontrado pelo codigo")

        if job.entity == "os_types":
            if row_data.get("client_cnpj") or row_data.get("client_code"):
                if not _resolve_client_id(db, job.tenant_id, row_data):
                    _append_error(errors, idx, "client_cnpj", "Cliente nao encontrado pelo CNPJ/Codigo")

        if not any(e["row_number"] == idx for e in errors):
            exists = _detect_existing(db, job.entity, job.tenant_id, row_data)
            if _should_skip(job.mode, bool(exists)):
                preview["skipped"] += 1
            elif exists:
                preview["updated"] += 1
            else:
                preview["created"] += 1
            if len(preview["samples"]) < MAX_PREVIEW_ROWS:
                preview["samples"].append(row_data)

    preview["errors"] = len(errors)

    if errors:
        db.bulk_insert_mappings(models.ImportRowError, [
            {**e, "import_job_id": job.id} for e in errors
        ])

    job.preview_json = preview
    job.summary_json = {
        "created": preview["created"],
        "updated": preview["updated"],
        "skipped": preview["skipped"],
        "errors_count": preview["errors"],
        "warnings_count": 0,
    }
    job.status = "ready_to_confirm" if preview["errors"] == 0 else "failed"
    db.commit()

    if errors:
        report_url = generate_error_report(storage, job, header, canonical_headers, errors)
        job.error_report_url = report_url
        db.commit()

    return preview


def _detect_existing(db, entity: str, tenant_id: str, row_data: Dict[str, object]):
    if entity == "employees":
        email = row_data.get("email")
        if not email:
            return None
        return (
            db.query(models.Colaborador)
            .filter(models.Colaborador.tenant_id == tenant_id, models.Colaborador.email == email)
            .first()
        )
    if entity == "clients":
        doc = row_data.get("document")
        code = row_data.get("client_code")
        query = db.query(models.Client).filter(models.Client.tenant_id == tenant_id)
        if doc:
            query = query.filter(models.Client.document == doc)
        elif code:
            query = query.filter(models.Client.client_code == code)
        else:
            return None
        return query.first()
    if entity == "sites":
        code = row_data.get("site_code")
        if not code:
            return None
        return (
            db.query(models.Site)
            .filter(models.Site.tenant_id == tenant_id, models.Site.code == code)
            .first()
        )
    if entity == "assets":
        tag = row_data.get("tag")
        if not tag:
            return None
        return (
            db.query(models.Asset)
            .filter(models.Asset.tenant_id == tenant_id, models.Asset.tag == tag)
            .first()
        )
    if entity == "os_types":
        name = row_data.get("name")
        client_id = _resolve_client_id(db, tenant_id, row_data)
        query = db.query(models.OSType).filter(models.OSType.tenant_id == tenant_id, models.OSType.name == name)
        if client_id:
            query = query.filter(models.OSType.client_id == client_id)
        return query.first()
    if entity == "questionnaires":
        title = row_data.get("title")
        version = row_data.get("version") or 1
        return (
            db.query(models.Questionnaire)
            .filter(
                models.Questionnaire.tenant_id == tenant_id,
                models.Questionnaire.title == title,
                models.Questionnaire.version == version,
            )
            .first()
        )
    return None


def _resolve_client_id(db, tenant_id: str, row_data: Dict[str, object]):
    cnpj = row_data.get("client_cnpj")
    code = row_data.get("client_code")
    if not cnpj and not code:
        return None
    query = db.query(models.Client).filter(models.Client.tenant_id == tenant_id)
    if cnpj:
        query = query.filter(models.Client.document == cnpj)
    if code:
        query = query.filter(models.Client.client_code == code)
    client = query.first()
    if client:
        return client.id
    return None


def _resolve_customer_account_id(db, tenant_id: str, row_data: Dict[str, object]):
    cnpj = row_data.get("customer_account_cnpj")
    name = row_data.get("customer_account_name")
    query = db.query(models.CustomerAccount).filter(models.CustomerAccount.tenant_id == tenant_id)
    if cnpj:
        query = query.filter(models.CustomerAccount.cnpj == cnpj)
    if name:
        query = query.filter(models.CustomerAccount.name == name)
    account = query.first()
    return account.id if account else None


def run_job(db, job: models.ImportJob) -> dict:
    config = ENTITY_CONFIGS.get(job.entity)
    if not config:
        raise ImportValidationError("Entidade nao suportada.")

    if job.entity == "questionnaires":
        return _run_questionnaire_job(db, job)

    if job.status not in {"queued", "running"}:
        return job.summary_json or {}

    storage = StorageClient()
    local_path = storage.download_to_temp(job.file_url)
    header, rows = iter_rows(local_path)
    header_map = make_header_map(config.template_columns)
    canonical_headers = [header_map.get(normalize_header(h), "") for h in header]

    created = updated = skipped = 0
    errors = []
    buffer: List[Dict[str, object]] = []

    job.status = "running"
    if not job.started_at:
        job.started_at = datetime.utcnow()
    db.commit()

    for idx, row in enumerate(rows, start=3):
        row_data: Dict[str, object] = {}
        for key, value in zip(canonical_headers, row):
            if not key:
                continue
            raw = value.strip() if isinstance(value, str) else value
            if raw in (None, ""):
                continue
            transformer = config.transformers.get(key)
            try:
                row_data[key] = transformer(str(raw)) if transformer else raw
            except Exception:
                _append_error(errors, idx, key, f"Valor invalido para {key}")
        if not any(str(cell).strip() for cell in row):
            continue
        buffer.append(row_data)
        if len(buffer) >= CHUNK_SIZE:
            c, u, s = _apply_chunk(db, job, buffer)
            created += c
            updated += u
            skipped += s
            buffer = []

    if buffer:
        c, u, s = _apply_chunk(db, job, buffer)
        created += c
        updated += u
        skipped += s

    job.status = "completed"
    job.finished_at = datetime.utcnow()
    job.summary_json = {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors_count": len(errors),
        "warnings_count": 0,
    }
    _record_audit(
        db,
        job,
        "bulk.import.completed",
        {"created": created, "updated": updated, "skipped": skipped, "errors": len(errors)},
    )
    db.commit()
    return job.summary_json


def _run_questionnaire_job(db, job: models.ImportJob) -> dict:
    storage = StorageClient()
    local_path = storage.download_to_temp(job.file_url)
    header, rows = iter_rows(local_path)
    header_map = make_header_map(ENTITY_CONFIGS[job.entity].template_columns)
    canonical_headers = [header_map.get(normalize_header(h), "") for h in header]

    grouped: Dict[Tuple[str, int], List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        if not any(str(cell).strip() for cell in row):
            continue
        row_data: Dict[str, object] = {}
        for key, value in zip(canonical_headers, row):
            if not key:
                continue
            raw = value.strip() if isinstance(value, str) else value
            if raw in (None, ""):
                continue
            transformer = ENTITY_CONFIGS[job.entity].transformers.get(key)
            row_data[key] = transformer(str(raw)) if transformer else raw
        title = row_data.get("title")
        version = row_data.get("version") or 1
        if not title:
            continue
        grouped[(title, int(version))].append(row_data)

    job.status = "running"
    if not job.started_at:
        job.started_at = datetime.utcnow()
    db.commit()

    created = updated = skipped = 0
    for (title, version), items in grouped.items():
        questionnaire = (
            db.query(models.Questionnaire)
            .filter(
                models.Questionnaire.tenant_id == job.tenant_id,
                models.Questionnaire.title == title,
                models.Questionnaire.version == version,
            )
            .first()
        )
        if questionnaire and job.mode == "create_only":
            skipped += 1
            continue
        if not questionnaire and job.mode == "update_only":
            skipped += 1
            continue
        if questionnaire:
            questionnaire.updated_at = datetime.utcnow()
            updated += 1
        else:
            questionnaire = models.Questionnaire(
                tenant_id=job.tenant_id,
                title=title,
                version=version,
                status="ATIVO",
            )
            db.add(questionnaire)
            db.flush()
            created += 1

        db.query(models.QuestionnaireItem).filter(
            models.QuestionnaireItem.questionnaire_id == questionnaire.id
        ).delete()

        for index, row in enumerate(items):
            db.add(
                models.QuestionnaireItem(
                    questionnaire_id=questionnaire.id,
                    question_text=row.get("question_text"),
                    required=bool(row.get("required")),
                    answer_type=row.get("answer_type"),
                    items=row.get("items") or [],
                    order_index=index,
                )
            )
        db.commit()

    job.status = "completed"
    job.finished_at = datetime.utcnow()
    job.summary_json = {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors_count": 0,
        "warnings_count": 0,
    }
    _record_audit(
        db,
        job,
        "bulk.import.completed",
        {"created": created, "updated": updated, "skipped": skipped, "errors": 0},
    )
    db.commit()
    return job.summary_json


def _apply_chunk(db, job: models.ImportJob, rows: List[Dict[str, object]]):
    created = updated = skipped = 0
    for row in rows:
        existing = _detect_existing(db, job.entity, job.tenant_id, row)
        if _should_skip(job.mode, bool(existing)):
            skipped += 1
            continue
        if job.entity == "employees":
            if existing:
                existing.name = row.get("nome", existing.name)
                existing.funcao = row.get("funcao", existing.funcao)
                existing.status = row.get("status", existing.status)
                existing.telefone = row.get("telefone", existing.telefone)
                existing.contrato = row.get("contrato", existing.contrato)
                existing.unidade = row.get("unidade", existing.unidade)
                existing.coordenador_nome = row.get("coordenador_nome", existing.coordenador_nome)
                existing.supervisor_nome = row.get("supervisor_nome", existing.supervisor_nome)
                existing.especialidades = row.get("especialidades", existing.especialidades)
                existing.observacoes = row.get("observacoes", existing.observacoes)
                updated += 1
            else:
                db.add(
                    models.Colaborador(
                        tenant_id=job.tenant_id,
                        nome=row.get("nome"),
                        funcao=row.get("funcao"),
                        email=row.get("email"),
                        telefone=row.get("telefone"),
                        status=row.get("status") or "ATIVO",
                        contrato=row.get("contrato"),
                        unidade=row.get("unidade"),
                        coordenador_nome=row.get("coordenador_nome"),
                        supervisor_nome=row.get("supervisor_nome"),
                        especialidades=row.get("especialidades"),
                        observacoes=row.get("observacoes"),
                    )
                )
                created += 1
        elif job.entity == "clients":
            if existing:
                existing.name = row.get("name", existing.name)
                existing.client_code = row.get("client_code", existing.client_code)
                existing.document = row.get("document", existing.document)
                existing.contract = row.get("contract", existing.contract)
                existing.address = row.get("address", existing.address)
                existing.status = row.get("status", existing.status)
                updated += 1
            else:
                db.add(
                    models.Client(
                        tenant_id=job.tenant_id,
                        name=row.get("name"),
                        client_code=row.get("client_code"),
                        document=row.get("document"),
                        contract=row.get("contract"),
                        address=row.get("address"),
                        status=row.get("status") or "active",
                    )
                )
                created += 1
        elif job.entity == "sites":
            account_id = _resolve_customer_account_id(db, job.tenant_id, row)
            if existing:
                existing.name = row.get("name", existing.name)
                existing.status = row.get("status", existing.status)
                existing.address = row.get("address", existing.address)
                existing.customer_account_id = account_id or existing.customer_account_id
                updated += 1
            else:
                db.add(
                    models.Site(
                        tenant_id=job.tenant_id,
                        code=row.get("site_code"),
                        name=row.get("name"),
                        status=row.get("status") or "ATIVO",
                        address=row.get("address"),
                        customer_account_id=account_id,
                    )
                )
                created += 1
        elif job.entity == "assets":
            client_id = _resolve_client_id(db, job.tenant_id, row)
            site = None
            if row.get("site_code"):
                site = (
                    db.query(models.Site)
                    .filter(models.Site.tenant_id == job.tenant_id, models.Site.code == row.get("site_code"))
                    .first()
                )
            if existing:
                existing.name = row.get("name", existing.name)
                existing.asset_type = row.get("asset_type", existing.asset_type)
                existing.status = row.get("status", existing.status)
                existing.client_id = client_id or existing.client_id
                existing.site_id = site.id if site else existing.site_id
                updated += 1
            else:
                db.add(
                    models.Asset(
                        tenant_id=job.tenant_id,
                        tag=row.get("tag"),
                        name=row.get("name"),
                        asset_type=row.get("asset_type"),
                        status=row.get("status"),
                        client_id=client_id,
                        site_id=site.id if site else None,
                    )
                )
                created += 1
        elif job.entity == "os_types":
            client_id = _resolve_client_id(db, job.tenant_id, row)
            if existing:
                existing.description = row.get("description", existing.description)
                if row.get("is_active") is not None:
                    existing.is_active = row.get("is_active")
                existing.client_id = client_id or existing.client_id
                updated += 1
            else:
                db.add(
                    models.OSType(
                        tenant_id=job.tenant_id,
                        name=row.get("name"),
                        description=row.get("description"),
                        is_active=row.get("is_active") if row.get("is_active") is not None else True,
                        client_id=client_id,
                    )
                )
                created += 1
        elif job.entity == "questionnaires":
            title = row.get("title")
            version = row.get("version") or 1
            questionnaire = _detect_existing(db, job.entity, job.tenant_id, row)
            if questionnaire:
                questionnaire.title = title
                questionnaire.version = version
                questionnaire.updated_at = datetime.utcnow()
            else:
                questionnaire = models.Questionnaire(
                    tenant_id=job.tenant_id,
                    title=title,
                    version=version,
                    status="ATIVO",
                )
                db.add(questionnaire)
                db.flush()
                created += 1
            db.query(models.QuestionnaireItem).filter(
                models.QuestionnaireItem.questionnaire_id == questionnaire.id
            ).delete()
            items = row.get("items") or []
            db.add(
                models.QuestionnaireItem(
                    questionnaire_id=questionnaire.id,
                    question_text=row.get("question_text"),
                    required=bool(row.get("required")),
                    answer_type=row.get("answer_type"),
                    items=items,
                    order_index=0,
                )
            )
            updated += 1
        else:
            skipped += 1

    db.commit()
    return created, updated, skipped


def generate_error_report(storage: StorageClient, job: models.ImportJob, headers: List[str], canonical_headers: List[str], errors: List[dict]) -> str:
    error_map = defaultdict(list)
    for error in errors:
        error_map[error["row_number"]].append(error)

    wb = Workbook()
    ws = wb.active
    ws.title = "ERROS"
    ws.append(headers + ["__status", "__error_fields", "__messages"])

    local_path = storage.download_to_temp(job.file_url)
    _, rows = iter_rows(local_path)

    for idx, row in enumerate(rows, start=3):
        row_errors = error_map.get(idx, [])
        if not row_errors:
            continue
        fields = ";".join(sorted({err["field"] for err in row_errors if err["field"]}))
        messages = ";".join([err["message"] for err in row_errors])
        ws.append(row + ["ERRO", fields, messages])

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(tmp_path)
    with open(tmp_path, "rb") as handle:
        content = handle.read()
    dest = f"bulk/errors/{job.tenant_id}/{job.id}.xlsx"
    return storage.upload_bytes(content, dest, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _record_audit(db, job: models.ImportJob, action: str, payload: dict) -> None:
    db.add(
        models.AuditLog(
            tenant_id=job.tenant_id,
            user_id=job.created_by_user_id,
            action=action,
            resource_type="import_job",
            resource_id=job.id,
            payload_resumo=payload,
        )
    )
